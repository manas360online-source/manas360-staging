"""
MANAS360 Test Data Loader
─────────────────────────
Loads test data from Excel OR runs the SQL seed directly.

USAGE:
  # Option A: Pure SQL (no dependencies beyond psycopg2)
  python load_test_data.py --mode sql

  # Option B: From Excel file
  python load_test_data.py --mode excel --file MANAS360_Test_Data_Complete.xlsx

  # Option C: Reset & reload (drops all tables first)
  python load_test_data.py --mode sql --reset

  # Option D: Validate only (check counts after loading)
  python load_test_data.py --mode validate

REQUIREMENTS:
  pip install psycopg2-binary openpyxl

ENV VARS (or edit DB_CONFIG below):
  MANS360_DB_HOST=localhost
  MANS360_DB_PORT=5432
  MANS360_DB_NAME=mans360_dev
  MANS360_DB_USER=mans360admin
  MANS360_DB_PASSWORD=your_password
"""

import os
import sys
import argparse
from datetime import datetime

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed, rely on system env vars

# ══════════════════════════════════════
# DATABASE CONFIG — Edit these or use env vars
# ══════════════════════════════════════
DB_CONFIG = {
    'host': os.getenv('MANS360_DB_HOST', 'localhost'),
    'port': int(os.getenv('MANS360_DB_PORT', '5432')),
    'dbname': os.getenv('MANS360_DB_NAME', 'mans360_dev'),
    'user': os.getenv('MANS360_DB_USER', 'mans360admin'),
    'password': os.getenv('MANS360_DB_PASSWORD', 'changeme'),
}

SQL_SEED_FILE = 'mans360_test_seed.sql'
EXCEL_FILE = 'MANAS360_Test_Data_Complete.xlsx'


def get_connection():
    """Connect to PostgreSQL."""
    try:
        import psycopg2
        conn = psycopg2.connect(**DB_CONFIG)
        conn.autocommit = False
        print(f"✅ Connected to {DB_CONFIG['dbname']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}")
        return conn
    except ImportError:
        print("❌ psycopg2 not installed. Run: pip install psycopg2-binary")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Connection failed: {e}")
        print(f"   Config: {DB_CONFIG['user']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['dbname']}")
        print(f"   Set env vars or edit DB_CONFIG in this script.")
        sys.exit(1)


def load_sql_seed(reset=False):
    """Load data from the SQL seed file."""
    if not os.path.exists(SQL_SEED_FILE):
        print(f"❌ SQL seed file not found: {SQL_SEED_FILE}")
        print(f"   Make sure mans360_test_seed.sql is in the same directory.")
        sys.exit(1)
    
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        if reset:
            print("🗑️  Dropping existing tables...")
            cur.execute("""
                DROP TABLE IF EXISTS test_scenarios CASCADE;
                DROP TABLE IF EXISTS fact_qr_tracking CASCADE;
                DROP TABLE IF EXISTS doctor_asset_orders CASCADE;
                DROP TABLE IF EXISTS doctor_redemptions CASCADE;
                DROP TABLE IF EXISTS doctor_credits CASCADE;
                DROP TABLE IF EXISTS doctor_referrals CASCADE;
                DROP TABLE IF EXISTS doctor_profiles CASCADE;
                DROP TABLE IF EXISTS audit_logs CASCADE;
                DROP TABLE IF EXISTS wallet_transactions CASCADE;
                DROP TABLE IF EXISTS leads CASCADE;
                DROP TABLE IF EXISTS certificates CASCADE;
                DROP TABLE IF EXISTS assessments CASCADE;
                DROP TABLE IF EXISTS cbt_session_data CASCADE;
                DROP TABLE IF EXISTS patient_profiles CASCADE;
                DROP TABLE IF EXISTS wallets CASCADE;
                DROP TABLE IF EXISTS sessions CASCADE;
                DROP TABLE IF EXISTS provider_credentials CASCADE;
                DROP TABLE IF EXISTS subscriptions CASCADE;
                DROP TABLE IF EXISTS iam_role_permissions CASCADE;
                DROP TABLE IF EXISTS permissions CASCADE;
                DROP TABLE IF EXISTS users CASCADE;
            """)
            # Drop all enums
            for enum in ['user_role','subscription_tier','subscription_status',
                         'payment_status','session_type','session_mode',
                         'phq9_severity','gad7_severity','payout_status',
                         'profile_status','verification_status']:
                cur.execute(f"DROP TYPE IF EXISTS {enum} CASCADE;")
            conn.commit()
            print("✅ Tables dropped.")
        
        print(f"📄 Loading {SQL_SEED_FILE}...")
        with open(SQL_SEED_FILE, 'r', encoding='utf-8') as f:
            sql = f.read()
        
        cur.execute(sql)
        conn.commit()
        print("✅ SQL seed loaded successfully!")
        
        # Print counts
        validate_data(conn)
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Error: {e}")
        sys.exit(1)
    finally:
        cur.close()
        conn.close()


def load_from_excel(filepath=None):
    """Load data from Excel file into PostgreSQL."""
    try:
        import openpyxl
    except ImportError:
        print("❌ openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)
    
    filepath = filepath or EXCEL_FILE
    if not os.path.exists(filepath):
        print(f"❌ Excel file not found: {filepath}")
        sys.exit(1)
    
    print(f"📊 Reading {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        # ── Load Users ──
        ws = wb['Users_Master']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"   Users: {len(rows)} rows")
        
        for r in rows:
            if not r[0]: continue
            langs = r[11].split(',') if r[11] else []
            cur.execute("""
                INSERT INTO users (id, role, first_name, last_name, email, phone,
                    gender, age, city, state, language_primary, languages_all,
                    aadhaar_last4, pan_masked, created_at, last_login,
                    is_active, is_verified, profile_complete_pct, notes)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                    last_login = EXCLUDED.last_login,
                    is_active = EXCLUDED.is_active,
                    notes = EXCLUDED.notes
            """, (
                r[0], r[1], r[2], r[3], r[4], r[5],
                r[6], r[7], r[8], r[9], r[10], langs,
                r[12], r[13], r[14], r[15],
                r[16], r[17], r[18], r[19]
            ))
        
        # ── Load Subscriptions ──
        ws = wb['Subscriptions']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"   Subscriptions: {len(rows)} rows")
        
        for r in rows:
            if not r[0]: continue
            cur.execute("""
                INSERT INTO subscriptions (user_id, tier, status, plan_type,
                    plan_price_monthly, trial_started_at, trial_ends_at,
                    premium_started_at, premium_ends_at, payment_method,
                    payment_status, auto_renew, last_payment_date,
                    next_billing_date, days_remaining, renewal_reminders_sent,
                    corporate_org, festival_free_access, promo_code_used, notes)
                VALUES (%s,%s,%s,%s,%s,
                    NULLIF(%s,'')::timestamp, NULLIF(%s,'')::timestamp,
                    NULLIF(%s,'')::timestamp, NULLIF(%s,'')::timestamp,
                    %s,%s,%s,
                    NULLIF(%s,'')::timestamp, NULLIF(%s,'')::timestamp,
                    %s,%s,%s,%s,%s,%s)
                ON CONFLICT (user_id) DO UPDATE SET
                    tier = EXCLUDED.tier,
                    status = EXCLUDED.status,
                    days_remaining = EXCLUDED.days_remaining,
                    notes = EXCLUDED.notes
            """, (
                r[0], r[3], r[4], r[9],
                r[10], r[5] or '', r[6] or '',
                r[7] or '', r[8] or '',
                r[11], r[12] or 'unpaid', r[13],
                r[14] or '', r[15] or '',
                r[16], r[17], r[18],
                True if str(r[19]).lower() == 'true' else False,
                r[20], r[21]
            ))
        
        # ── Load Wallets ──
        ws = wb['Wallet_Credits']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"   Wallets: {len(rows)} rows")
        
        for r in rows:
            if not r[0]: continue
            cur.execute("""
                INSERT INTO wallets (user_id, balance, total_earned, total_withdrawn,
                    total_spent_leads, pending_payout, last_credit_date, last_debit_date,
                    last_payout_date, payout_bank, payout_status, min_payout_threshold,
                    leads_purchased_total, leads_purchased_this_month,
                    session_credits_remaining, bonus_credits, referral_credits,
                    festival_bonus, notes)
                VALUES (%s,%s,%s,%s,%s,%s,
                    NULLIF(%s,'')::timestamp, NULLIF(%s,'')::timestamp,
                    NULLIF(%s,'')::timestamp, %s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (user_id) DO UPDATE SET
                    balance = EXCLUDED.balance,
                    total_earned = EXCLUDED.total_earned,
                    notes = EXCLUDED.notes
            """, (
                r[0], r[3], r[4], r[5], r[6], r[7],
                r[8] or '', r[9] or '', r[10] or '',
                r[11], r[12] or 'N/A', r[13],
                r[14], r[15], r[16], r[17], r[18], r[19], r[20]
            ))
        
        # ── Load Sessions ──
        ws = wb['Sessions_BeforeAfter']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        print(f"   Sessions: {len(rows)} rows")
        
        for r in rows:
            if not r[0]: continue
            # Map provider_role for IVR
            p_role = r[5] if r[5] != 'system' else 'therapist'
            cur.execute("""
                INSERT INTO sessions (id, patient_id, patient_name, provider_id,
                    provider_name, provider_role, session_date, session_time,
                    duration_min, session_type, session_mode, session_number,
                    language, pre_phq9_score, pre_phq9_severity, pre_gad7_score,
                    pre_gad7_severity, pre_mood_self_rated, pre_sleep_hours,
                    pre_energy_level, post_phq9_score, post_phq9_severity,
                    post_gad7_score, post_gad7_severity, post_mood_self_rated,
                    post_sleep_hours, post_energy_level,
                    session_rating_patient, session_rating_provider,
                    session_fee, fee_payment_status, provider_payout,
                    notes_summary, crisis_flag, follow_up_scheduled)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                        %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                        %s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO NOTHING
            """, (
                r[0], r[1], r[2], r[3], r[4], p_role,
                r[6], r[7], r[8], r[9], r[10], r[11], r[12],
                r[13], r[14], r[15], r[16], r[17], r[18], r[19],
                r[20], r[21], r[22], r[23], r[24], r[25], r[26],
                # skip phq9_delta(27), gad7_delta(28), mood_delta(29) — generated columns
                r[30], r[31], r[32],
                r[33] if r[33] else 'unpaid',
                r[34] if r[34] and r[34] != 'N/A' else 0,
                r[35], r[36], r[37]
            ))
        
        conn.commit()
        print("✅ Excel data loaded successfully!")
        validate_data(conn)
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Error loading Excel: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        cur.close()
        conn.close()
        wb.close()


def validate_data(conn=None):
    """Print counts from all tables to verify data."""
    close_conn = False
    if conn is None:
        conn = get_connection()
        close_conn = True
    
    cur = conn.cursor()
    print("\n" + "═" * 50)
    print("📊 DATA VALIDATION")
    print("═" * 50)
    
    queries = [
        ("Users by role", "SELECT role, COUNT(*) FROM users GROUP BY role ORDER BY role"),
        ("Subscriptions by tier/status", "SELECT tier, status, COUNT(*) FROM subscriptions GROUP BY tier, status ORDER BY tier, status"),
        ("Wallets", "SELECT COUNT(*), SUM(balance), SUM(total_earned) FROM wallets"),
        ("Sessions by type", "SELECT session_type, COUNT(*), AVG(phq9_delta)::numeric(4,1) AS avg_phq9_change FROM sessions GROUP BY session_type"),
        ("Crisis sessions", "SELECT COUNT(*) FROM sessions WHERE crisis_flag = TRUE"),
        ("Provider credentials", "SELECT profile_status, COUNT(*) FROM provider_credentials GROUP BY profile_status"),
        ("Patient Profiles", "SELECT COUNT(*) FROM patient_profiles"),
        ("CBT Data", "SELECT COUNT(*), worksheet_type FROM cbt_session_data GROUP BY worksheet_type"),
        ("Assessments", "SELECT assessment_type, COUNT(*), AVG(score)::numeric(4,1) FROM assessments GROUP BY assessment_type"),
        ("Certificates", "SELECT certificate_type, COUNT(*) FROM certificates GROUP BY certificate_type"),
        ("Leads", "SELECT status, COUNT(*) FROM leads GROUP BY status"),
        ("Wallet Transactions", "SELECT transaction_type, COUNT(*), SUM(amount) FROM wallet_transactions GROUP BY transaction_type"),
        ("Audit Logs", "SELECT action, COUNT(*) FROM audit_logs GROUP BY action"),
        ("Doctor Profiles", "SELECT referral_tier, COUNT(*) FROM doctor_profiles GROUP BY referral_tier"),
        ("Doctor Referrals", "SELECT status, COUNT(*) FROM doctor_referrals GROUP BY status"),
        ("Doctor Dashboard View", "SELECT * FROM v_doctor_dashboard"),
    ]
    
    for label, query in queries:
        try:
            cur.execute(query)
            rows = cur.fetchall()
            print(f"\n  {label}:")
            for row in rows:
                # Format output nicely:
                # 1. Single value -> print directly
                # 2. Multiple values -> join with |
                formatted_values = []
                for val in row:
                    if isinstance(val, (int, float)):
                        formatted_values.append(str(val))
                    elif isinstance(val, str):
                        formatted_values.append(val)
                    else:
                        formatted_values.append(str(val))
                
                print(f"    {' | '.join(formatted_values)}")
        except Exception as e:
            print(f"\n  {label}: ⚠️  {e}")
    
    # Test the views
    for view in ['v_patient_360', 'v_provider_dashboard', 'v_session_outcomes']:
        try:
            cur.execute(f"SELECT COUNT(*) FROM {view}")
            count = cur.fetchone()[0]
            print(f"\n  View {view}: {count} rows ✅")
        except Exception as e:
            print(f"\n  View {view}: ⚠️  {e}")
    
    print("\n" + "═" * 50)
    cur.close()
    if close_conn:
        conn.close()


def generate_env_template():
    """Create a .env template file."""
    env_content = """# MANAS360 Test Database Configuration
# Copy this to .env and update values

MANS360_DB_HOST=localhost
MANS360_DB_PORT=5432
MANS360_DB_NAME=mans360_dev
MANS360_DB_USER=mans360admin
MANS360_DB_PASSWORD=changeme

# For AWS Lightsail:
# MANS360_DB_HOST=ls-xxxxx.ap-south-1.rds.amazonaws.com
# MANS360_DB_PASSWORD=your_rds_password

# For Supabase:
# MANS360_DB_HOST=db.xxxxx.supabase.co
# MANS360_DB_PORT=5432
# MANS360_DB_NAME=postgres
# MANS360_DB_USER=postgres
# MANS360_DB_PASSWORD=your_supabase_password
"""
    with open('.env.template', 'w') as f:
        f.write(env_content)
    print("✅ Created .env.template — copy to .env and update values.")


def main():
    parser = argparse.ArgumentParser(
        description='MANAS360 Test Data Loader',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
EXAMPLES:
  python load_test_data.py --mode sql              # Load from SQL seed
  python load_test_data.py --mode sql --reset      # Drop + reload everything
  python load_test_data.py --mode excel            # Load from Excel file  
  python load_test_data.py --mode validate         # Check data counts
  python load_test_data.py --mode env              # Generate .env template
        """
    )
    parser.add_argument('--mode', choices=['sql', 'excel', 'validate', 'env'],
                        default='sql', help='Loading mode (default: sql)')
    parser.add_argument('--file', default=EXCEL_FILE,
                        help=f'Excel file path (default: {EXCEL_FILE})')
    parser.add_argument('--reset', action='store_true',
                        help='Drop all tables before loading')
    
    args = parser.parse_args()
    
    print("=" * 50)
    print("🧠 MANAS360 Test Data Loader")
    print(f"   Mode: {args.mode}")
    print(f"   Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 50)
    
    if args.mode == 'env':
        generate_env_template()
    elif args.mode == 'sql':
        load_sql_seed(reset=args.reset)
    elif args.mode == 'excel':
        load_from_excel(args.file)
    elif args.mode == 'validate':
        validate_data()


if __name__ == '__main__':
    main()

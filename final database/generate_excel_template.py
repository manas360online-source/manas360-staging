
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_NAME = "MANAS360_Test_Data_Complete.xlsx"

def create_excel_template():
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # 1. Users_Master Sheet
    # ---------------------------------------------------------
    ws_users = wb.active
    ws_users.title = "Users_Master"
    
    user_headers = [
        "ID", "Role", "First Name", "Last Name", "Email", "Phone",
        "Gender", "Age", "City", "State", "Primary Language", "All Languages (comma-sep)",
        "Aadhaar (Last 4)", "PAN (Masked)", "Created At", "Last Login",
        "Is Active", "Is Verified", "Profile Complete %", "Notes"
    ]
    
    user_data = [
        ["USR-PAT-001", "patient", "Priya", "Sharma", "priya.sharma@gmail.com", "+919876543210", "Female", 28, "Bengaluru", "Karnataka", "Kannada", "Kannada,English,Hindi", "4521", "ABCPS****K", "2025-11-19", "2026-02-17 08:00", True, True, 100, "Premium active. 12 sessions completed."],
        ["USR-THR-001", "therapist", "Dr. Lakshmi", "Iyer", "lakshmi.iyer@gmail.com", "+919801234567", "Female", 38, "Bengaluru", "Karnataka", "English", "English,Kannada,Tamil", "3456", "KLMNP****U", "2025-08-17", "2026-02-17 09:00", True, True, 100, "Top rated. CBT specialist."],
    ]
    
    setup_sheet(ws_users, user_headers, user_data)

    # ---------------------------------------------------------
    # 2. Subscriptions Sheet
    # ---------------------------------------------------------
    ws_subs = wb.create_sheet("Subscriptions")
    
    sub_headers = [
        "User ID", "User Name (Ref)", "User Role (Ref)", "Tier", "Status",
        "Trial Start", "Trial End", "Premium Start", "Premium End",
        "Plan Type", "Monthly Price", "Payment Method", "Payment Status",
        "Auto Renew", "Last Payment", "Next Billing", "Days Remaining",
        "Reminders Sent", "Corporate Org", "Festival Free", "Promo Code", "Notes"
    ]
    
    sub_data = [
        ["USR-PAT-001", "Priya Sharma", "patient", "premium", "active", "2025-11-19", "2025-12-19", "2025-12-19", "2026-03-02", "monthly", 299, "UPI (PhonePe)", "paid", True, "2026-01-17", "2026-03-02", 13, 0, "", False, "WELCOME50", "Auto-renewing"],
    ]
    
    setup_sheet(ws_subs, sub_headers, sub_data)

    # ---------------------------------------------------------
    # 3. Wallet_Credits Sheet
    # ---------------------------------------------------------
    ws_wallets = wb.create_sheet("Wallet_Credits")
    
    wallet_headers = [
        "User ID", "User Name (Ref)", "User Role (Ref)", "Balance",
        "Total Earned", "Total Withdrawn", "Total Spent (Leads)", "Pending Payout",
        "Last Credit Date", "Last Debit Date", "Last Payout Date",
        "Payout Bank", "Payout Status", "Min Payout Threshold",
        "Leads Purchased Total", "Leads Purchased This Month",
        "Session Credits Remaining", "Bonus Credits", "Referral Credits",
        "Festival Bonus", "Notes"
    ]
    
    wallet_data = [
        ["USR-PAT-001", "Priya Sharma", "patient", 150, 0, 0, 0, 0, "2026-02-10", "2026-02-15", "", "", "N/A", 0, 0, 0, 2, 0, 150, 0, "Referral credit"],
        ["USR-THR-001", "Dr. Lakshmi", "therapist", 18500, 85000, 62000, 4500, 3500, "2026-02-16", "2026-02-14", "2026-02-10", "HDFC ****4521", "processed", 500, 35, 8, 0, 0, 0, 0, "Top earner"],
    ]
    
    setup_sheet(ws_wallets, wallet_headers, wallet_data)

    # ---------------------------------------------------------
    # 4. Sessions_BeforeAfter Sheet
    # ---------------------------------------------------------
    ws_sessions = wb.create_sheet("Sessions_BeforeAfter")
    
    session_headers = [
        "Session ID", "Patient ID", "Patient Name", "Provider ID", "Provider Name", "Provider Role",
        "Date", "Time", "Duration (min)", "Type", "Mode", "Session #", "Language",
        "Pre PHQ-9", "Pre PHQ-9 Severity", "Pre GAD-7", "Pre GAD-7 Severity",
        "Pre Mood (0-10)", "Pre Sleep (hrs)", "Pre Energy (0-10)",
        "Post PHQ-9", "Post PHQ-9 Severity", "Post GAD-7", "Post GAD-7 Severity",
        "Post Mood (0-10)", "Post Sleep (hrs)", "Post Energy (0-10)",
        "PHQ9 Delta (Calc)", "GAD7 Delta (Calc)", "Mood Delta (Calc)",
        "Rating (Patient)", "Rating (Provider)", "Fee", "Fee Status",
        "Provider Payout", "Notes Summary", "Crisis Flag", "Follow-up Scheduled"
    ]
    
    session_data = [
        ["SES-001", "USR-PAT-001", "Priya Sharma", "USR-THR-001", "Dr. Lakshmi", "therapist", "2025-12-02", "10:00", 50, "individual", "video", 1, "English", 18, "mod_severe", 15, "severe", 3, 4, 2, 18, "mod_severe", 14, "moderate", 4, 4, 3, 0, -1, 1, 4, 4, 800, "paid", 680, "Initial intake", False, True],
    ]
    
    setup_sheet(ws_sessions, session_headers, session_data)

    # Save
    wb.save(FILE_NAME)
    print(f"✅ Created Excel template: {FILE_NAME}")


def setup_sheet(ws, headers, data):
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
        # Adjust column width approx based on header length
        ws.column_dimensions[get_column_letter(col_num)].width = len(header) + 5

    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.border = thin_border
            
            # Center align boolean and numbers typically
            if isinstance(cell_value, (int, bool)) and col_num > 1:
                cell.alignment = center_align

if __name__ == "__main__":
    create_excel_template()
